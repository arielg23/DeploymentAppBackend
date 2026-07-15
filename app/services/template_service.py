import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from models.site import Site

REQUIRED_HEADERS = ["unit_id", "unit_name", "sequence"]
OPTIONAL_HEADERS = ["barcode", "customer_name", "customer_id", "site_name"]
ALL_HEADERS = REQUIRED_HEADERS + OPTIONAL_HEADERS

EXAMPLE_ROW = ["UNIT16538", "Unit 1", 1, "AA000034", "Acme Corp", "CUST-001", "Acme HQ"]

REQUIRED_FILL = PatternFill(start_color="D9E8FF", end_color="D9E8FF", fill_type="solid")
HEADER_FONT = Font(bold=True)


def generate_unit_template(sites: list[Site]) -> bytes:
    """Build a blank unit-upload workbook: header row (required columns
    highlighted) plus one example row. If sites are given, the site_name
    column gets a dropdown restricted to known site names so a typo can't
    silently defeat the optional site-match check in the parser."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Units"

    for col_idx, header in enumerate(ALL_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        if header in REQUIRED_HEADERS:
            cell.fill = REQUIRED_FILL
        ws.column_dimensions[cell.column_letter].width = max(14, len(header) + 4)

    for col_idx, value in enumerate(EXAMPLE_ROW, start=1):
        ws.cell(row=2, column=col_idx, value=value)

    if sites:
        ref = wb.create_sheet("Sites (reference)")
        ref.append(["site_name", "site_id"])
        for site in sites:
            ref.append([site.site_name, site.site_id])
        ref.sheet_state = "hidden"

        site_name_col_letter = ws.cell(row=1, column=ALL_HEADERS.index("site_name") + 1).column_letter
        dv = DataValidation(
            type="list",
            formula1=f"'Sites (reference)'!$A$2:$A${len(sites) + 1}",
            allow_blank=True,
        )
        ws.add_data_validation(dv)
        dv.add(f"{site_name_col_letter}2:{site_name_col_letter}1000")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
