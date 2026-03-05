import uuid
from typing import BinaryIO

from fastapi import HTTPException, status
from sqlalchemy.ext.asyncio import AsyncSession

from models.site import Site
from models.unit import Unit
from models.upload import Upload, UploadStatus

REQUIRED_COLUMNS = {"site_id", "unit_id", "unit_name", "sequence"}


async def parse_and_create_upload(
    db: AsyncSession,
    file: BinaryIO,
    site_id: str,
    uploaded_by: uuid.UUID,
) -> dict:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
        ws = wb.active
    except Exception as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Cannot read Excel file: {e}")

    headers = [str(cell.value).strip().lower() if cell.value else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    missing = REQUIRED_COLUMNS - set(headers)
    if missing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Missing required columns: {', '.join(missing)}",
        )

    col = {name: idx for idx, name in enumerate(headers)}
    units_data = []
    errors = []
    seen_unit_ids = set()

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row_site_id = str(row[col["site_id"]]).strip() if row[col["site_id"]] else ""
        row_unit_id = str(row[col["unit_id"]]).strip() if row[col["unit_id"]] else ""
        unit_name = str(row[col["unit_name"]]).strip() if row[col["unit_name"]] else ""
        sequence = row[col["sequence"]]

        if not row_unit_id or not unit_name:
            errors.append(f"Row {row_idx}: missing unit_id or unit_name")
            continue
        if row_site_id and row_site_id != site_id:
            errors.append(f"Row {row_idx}: site_id mismatch (expected {site_id}, got {row_site_id})")
            continue
        if row_unit_id in seen_unit_ids:
            errors.append(f"Row {row_idx}: duplicate unit_id {row_unit_id}")
            continue

        try:
            sequence = int(sequence)
        except (TypeError, ValueError):
            errors.append(f"Row {row_idx}: invalid sequence value")
            continue

        seen_unit_ids.add(row_unit_id)
        units_data.append({
            "unit_id": row_unit_id,
            "unit_name": unit_name,
            "sequence": sequence,
            "customer_name": str(row[col["customer_name"]]).strip() if "customer_name" in col and row[col["customer_name"]] else None,
            "customer_id": str(row[col["customer_id"]]).strip() if "customer_id" in col and row[col["customer_id"]] else None,
        })

    if errors:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail={"errors": errors})

    if not units_data:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="No valid units found in file")

    # Ensure site exists
    from sqlalchemy import select
    site_result = await db.execute(select(Site).where(Site.site_id == site_id))
    site = site_result.scalar_one_or_none()
    if not site:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=f"Site {site_id} not found")

    upload = Upload(site_id=site_id, status=UploadStatus.INACTIVE, uploaded_by=uploaded_by)
    db.add(upload)
    await db.flush()

    for u in units_data:
        db.add(Unit(upload_id=upload.upload_id, site_id=site_id, **u))

    await db.commit()
    await db.refresh(upload)
    return {"upload_id": str(upload.upload_id), "row_count": len(units_data), "errors": []}
